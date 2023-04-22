import numpy
from openpyxl import load_workbook
import math
from scipy.optimize import fsolve
from scipy.stats import linregress
from dataclasses import dataclass
import numpy as np
import json
import csv
import hydra
from omegaconf import DictConfig

# defining constants
pfeed = 1.01
qsweep = 50
lmembrane = 0.1
dmembrane = 2.7


@dataclass
class Components:
    gas: str
    bg: float
    raw: int
    p_ms_ar: float
    rs_0: float
    real: float
    i_0: float
    rs_i: int
    p_ms_i: float
    p_ms_tot: float
    p_p_i: float
    p_p_Ar_i: float
    q_p_i_q_ar_i: float
    q_sweep_Ar: int
    q_permeate: float
    p_feed: float
    d_membrane: float
    a_membrane: float
    permeance_in_m: float
    permeance_in_gpu: float
    permeance_in_mol: float
    l_membrane: float
    permeance_in_barrer: float
    diffusion_coefficient: numpy.float64


def calc_avg_bg(sheet, gas_column):
    sum_bg_n = 0
    for k in range(2, sheet.max_row + 1):
        if sheet["H" + str(k)].value == "switch time":
            count = k - 1
            break
    for k in range(count - 30, count + 1):
        sum_bg_n = sum_bg_n + int(sheet[gas_column + str(k)].value)
    avg_bg_n = sum_bg_n / 31
    return avg_bg_n


def print_result(components):
    print(f"Gas: {components.gas}")
    print(f"BG: {components.bg:}")
    print(f"raw: {components.raw}")
    print(f"p-ms-ar:  {components.p_ms_ar}")
    print(f"rs-0: {components.rs_0}")
    print(f"Real: {components.real}")
    print(f"I-0: {components.i_0}")
    print(f"rs-i: {components.rs_i}")
    print(f"P-MS-I: {components.p_ms_i}")
    print(f"P-MS-Tot: {components.p_ms_tot}")
    print(f"P-p,i: {components.p_p_i}")
    print(f"P-p,Ar,i: {components.p_p_Ar_i}")
    print(f"Q-p,i/Q-Ar,i: {components.q_p_i_q_ar_i}")
    print(f"Q-Sweep(Ar): {components.q_sweep_Ar} mln/min")
    print(f"Q-permeate: {components.q_permeate} mln/min")
    print(f"P-feed: {components.p_feed} bar")
    print(f"d-membrane: {components.d_membrane} cm")
    print(f"A-membrane: {components.a_membrane} cm\u00b2")
    print(f"Permeance in m\u00b3(STP)/(m\u00b2 h bar): {components.permeance_in_m}")
    print(f"Permeance in GPU: {components.permeance_in_gpu}")
    print(f"Permeance in 10\u207B\u2077 mol/(m\u00b2 s Pa): {components.permeance_in_mol}")
    print(f"L-membrane in μm: {components.l_membrane}")
    print(f"Permeability in Barrer: {components.permeance_in_barrer}")
    print(f'Diffusion coefficient: {components.diffusion_coefficient} \n \n')


def calc_rsi(x, *data):
    io, pmsar, gas = data
    diction = {'CH4': [79.21, 233.34], 'N2': [-41.31, 91.58], 'He': [-58.92, 267.25]}
    if gas == 'CO2':
        return 366.542 * math.exp(0.18068 / (
                (pfeed * (io / x) / ((io / x) + pmsar)) / (pfeed * pmsar / ((io / x) + pmsar)) * 100 + 0.08308)) - x
    else:
        return diction[gas][0] * math.log10(
            (pfeed * (io / x) / ((io / x) + pmsar)) / (pfeed * pmsar / ((io / x) + pmsar)) * 100) \
            + diction[gas][1] - x


def calc_switch_row(sheet):
    for i in range(2, sheet.max_row + 1):
        if sheet["H" + str(i)].value == "switch time":
            return i


def computation(filename_f, rso_f, amembrane_f, gas_f, fileformat_f):
    workbook = load_workbook(filename=filename_f)
    sheet = workbook.active

    if gas_f == 'H2':
        gas_column = 'D'
    if gas_f == 'He':
        gas_column = 'C'
    if gas_f == 'CH4' or gas_f == 'N2':
        gas_column = 'E'
    if gas_f == 'CO2':
        gas_column = 'F'

    avg_bg = calc_avg_bg(sheet, gas_column)

    switch_row = calc_switch_row(sheet)
    raw = {}
    prms = {}
    sum_ = 0
    c = 0

    for j in range(switch_row, sheet.max_row + 1):
        if sheet[gas_column + str(j)].value in raw:
            raw[sheet[gas_column + str(j)].value] += 1
        else:
            if sheet[gas_column + str(j)].value != 0:
                raw[sheet[gas_column + str(j)].value] = 1
        if sheet["G" + str(j)].value in prms:
            prms[sheet["G" + str(j)].value] += 1
        else:
            prms[sheet["G" + str(j)].value] = 1
        sum_ = sum_ + sheet["G" + str(j)].value
        c = c + 1

    h2_raw = max(zip(raw.values(), raw.keys()))[1]
    h2_prms = max(zip(prms.values(), prms.keys()))[1]
    pmsar = sum_ / c
    real = h2_raw - avg_bg
    io = real * float(rso_f)
    data = (io, pmsar, gas_f)
    rsi = [1171] if gas_f == 'H2' else fsolve(calc_rsi, 1000, args=data)
    pmsi = io / rsi[0]
    pmstot = pmsar + pmsi
    ppi = 1.01 * pmsi / pmstot
    ppar = 1.01 * pmsar / pmstot
    qpiqar = ppi / ppar * 100
    qpermeate = qsweep * qpiqar / 100
    permeancebar = 0.6 * qpermeate / (pfeed - ppi) / float(amembrane_f)
    permeancegpu = 370 * permeancebar
    permeancepa = permeancegpu * 3.35 / 1000
    permeability = permeancegpu * lmembrane
    gas_list = [sheet[gas_column + str(i)].value for i in range(switch_row, sheet.max_row + 1)]
    processed_signal = [((gas_list[i] - avg_bg) * float(rso_f)) / rsi[0] for i in range(len(gas_list))]
    argon_list = [sheet["G" + str(i)].value for i in range(switch_row - 1, sheet.max_row + 1)]
    sum_argon_processed = [processed_signal[i] + argon_list[i] for i in range(len(processed_signal))]
    gas_in_bar = [1.01 * processed_signal[i] / sum_argon_processed[i] for i in range(len(processed_signal))]
    argon_in_bar = [1.01 * argon_list[i] / sum_argon_processed[i] for i in range(len(processed_signal))]
    gas_mln_min = [qsweep * gas_in_bar[i] / argon_in_bar[i] for i in range(len(gas_in_bar))]
    gas_cum_vol = [0]

    for i in range(1, len(gas_list)):
        gas_cum_vol.append(0.5 * (gas_mln_min[i] + gas_mln_min[i - 1]))
    gas_cum_vol_mln = [0]
    sum_gas_cum_vol = 0

    for i in range(1, len(gas_list)):
        sum_gas_cum_vol = sum_gas_cum_vol + gas_cum_vol[i]
        gas_cum_vol_mln.append(1 / 60 * sum_gas_cum_vol)

    x_axis = [i for i in range(len(gas_list))]
    slope, intercept, r_value, p_value, std_err = linregress(x_axis[23:], gas_cum_vol_mln[23:])
    diff_coefficient = -intercept / slope

    component = Components(gas_f, avg_bg, h2_raw, pmsar, float(rso_f), real, io, rsi[0], pmsi, pmstot, ppi, ppar,
                           qpiqar, qsweep, qpermeate, pfeed, dmembrane, float(amembrane_f), permeancebar,
                           permeancegpu, permeancepa, lmembrane, permeability, diff_coefficient)
    print_result(component)
    save_result(component, fileformat_f)


def save_result(component, fileformat):
    data = {"Gas": component.gas, "BG": component.bg, "Raw": component.raw, "p-ms-ar": component.p_ms_ar,
            "rs-0": component.rs_0, "Real": component.real, "I-O": component.i_0,
            "rs-i": str(component.rs_i), "P-MS-I": component.p_ms_i, "P-MS-Tot": component.p_ms_tot,
            "P-p,i": component.p_p_i, "P-p,Ar,i": component.p_p_Ar_i, "Q-p,i/Q-Ar,i": component.q_p_i_q_ar_i,
            "Q-Sweep(Ar)": component.q_sweep_Ar, "Q-permeate": component.q_permeate, "P-feed": component.p_feed,
            "d-membrane": component.d_membrane, "A-membrane": component.a_membrane,
            "Permeance in bar": component.permeance_in_m, "Permeance in GPU": component.permeance_in_gpu,
            "Permeance in Pa": component.permeance_in_mol, "L-membrane": component.l_membrane,
            "Permeability in Barrer": component.permeance_in_barrer,
            "Diffusion coefficient": component.diffusion_coefficient}

    if fileformat == "json":
        with open("output_in_json.json", "a") as outfile:
            json.dump(data, outfile)
    elif fileformat == "csv":
        fieldnames = ['Gas', 'BG', 'Raw', 'p-ms-ar', 'rs-0', 'Real', 'I-0', 'rs-i', 'P-MS-I', 'P-MS-Tot', 'P-p,i',
                      'P-p,Ar,i',
                      'Q-p,i/Q-Ar,i',
                      'Q-Sweep(Ar)', 'Q-permeate', 'P-feed', 'd-membrane', 'A-membrane', 'Permeance in bar',
                      'Permeance in GPU',
                      'Permeance in Pa', 'Permeability in Barrer', 'L-membrane', 'I-O', 'Diffusion coefficient']
        with open("output_in_csv.csv", 'a') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows([data])


@hydra.main(version_base=None, config_path="conf", config_name="config")
def my_app(cfg: DictConfig) -> None:
    computation(cfg['config1']['filename'], cfg['config1']['rso'], cfg['config1']['amembrane'], cfg['config1']['gas'], cfg['fileformat'])
    if cfg['config2']['filename'] is not None:
        computation(cfg['config2']['filename'], cfg['config2']['rso'], cfg['config2']['amembrane'], cfg['config2']['gas'], cfg['fileformat'])
    if cfg['config3']['filename'] is not None:
        computation(cfg['config3']['filename'], cfg['config3']['rso'], cfg['config3']['amembrane'], cfg['config3']['gas'], cfg['fileformat'])
    if cfg['config4']['filename'] is not None:
        computation(cfg['config4']['filename'], cfg['config4']['rso'], cfg['config4']['amembrane'], cfg['config4']['gas'], cfg['fileformat'])
    if cfg['config5']['filename'] is not None:
        computation(cfg['config5']['filename'], cfg['config5']['rso'], cfg['config5']['amembrane'], cfg['config5']['gas'], cfg['fileformat'])


if __name__ == "__main__":
    my_app()
